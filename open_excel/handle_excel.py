import openpyxl
import os
import sys

'''
步骤分析：
第一步：加载表格文件
第二步：获取所有的sheetname
第三步：获取第一个sheet
第四步：获取某个单元格的数据
第五步：获取最大的行数
第六步：获取最大的列数
第七步：加载某一列的数据

open_excel = openpyxl.load_workbook('/Users/guojun/Desktop/驾照数据表.xlsx')
sheet_name = open_excel.sheetnames
sheet01 = open_excel[sheet_name[0]]
value = sheet01.cell(1, 4).value
print(value)
row = sheet01.max_row
print('最大的行数:', row)
column = sheet01.max_column
print('最大的列数:', column)
excel_list = []
for i in sheet01[1]:
    excel_list.append(i.value)
    print(i)
print(excel_list)

'''


class HandExcel(object):

    def load_excel(self):
        '''
        1、加载数据
        注意：如需使用文件地址需要修改；不要用绝对路径，一定要用相对路径，便于部署，切记切记！
        :return:
        '''
        excel = openpyxl.load_workbook('/Users/guojun/Desktop/驾照数据表.xlsx')
        return excel

    def get_sheet(self, index):
        '''
        2、获取sheet表
        需要传参sheet表所在的索引。（索引从0开始，如果是第一个sheet，索引就是0，依次类推）
        :param index:
        :return:
        '''
        sheet_name = self.load_excel().sheetnames
        sheet = self.load_excel()[sheet_name[index]]
        return sheet

    def get_one_value(self, index, row, column):
        '''
        3、获取某个单元格的数据
        需要传参sheet表所在的索引、单元格所在的行与列
        注意：要用.value方法才能获取到值，不然返回的是一个sheet对象
        :param index:
        :param row:
        :param column:
        :return:
        '''
        one_value = self.get_sheet(index).cell(row, column).value
        return one_value

    def get_row(self, index):
        '''
        4、获取列表的最大行数
        后续使用excel时，需要调用此方法查看表格最大数据，避免索引越界引起崩溃
        :param index:
        :return:
        '''
        max_row = self.get_sheet(index).max_row
        return max_row

    def get_column(self, index):
        '''
        5、获取列表的最大列数
        :param index:
        :return:
        '''
        max_column = self.get_sheet(index).max_column
        return max_column

    def get_row_data(self, index, row):
        '''
        6、加载某一列的数据
        需要传参：sheet的索引、取出哪行的数据
        以列表的形式返回某行的数据，调用此方法后需自行遍历才能使用
        :param index:
        :param row:
        :return:
        '''
        row_list = []
        for i in self.get_sheet(index)[row]:
            row_list.append(i.value)
        return row_list


if __name__ == '__main__':
    hand_excel = HandExcel()
    hand_excel.get_sheet(0)
    print(hand_excel.get_row_data(0, 4))

