import openpyxl
import os
import sys

'''
open_excel = openpyxl.load_workbook('/Users/guojun/Desktop/驾照数据表.xlsx')
sheet_name = open_excel.sheetnames
sheet01 = open_excel[sheet_name[0]]
value = sheet01.cell(1, 4).value
print(value)
row = sheet01.max_row
print('最大的行数:', row)
column = sheet01.max_column
print('最大的列数:', column)
excel_list = []
for i in sheet01[1]:
    excel_list.append(i.value)
    print(i)
print(excel_list)


第一步：加载表格文件
第二步：获取所有的sheetname
第三步：获取第一个sheet
第四步：获取某个单元格的数据
第五步：获取最大的行数
第六步：获取最大的列数
第七步：加载某一列的数据

'''


class HandExcel(object):

    def load_excel(self):
        excel = openpyxl.load_workbook('/Users/guojun/Desktop/驾照数据表.xlsx')
        return excel

    def get_sheet(self, index):
        sheet_name = self.load_excel().sheetnames
        sheet = self.load_excel()[sheet_name[index]]
        return sheet

    def get_one_value(self, index, row, column):
        one_value = self.get_sheet(index).cell(row, column).value
        return one_value

    def get_row(self, index):
        max_row = self.get_sheet(index).max_row
        return max_row

    def get_column(self, index):
        max_column = self.get_sheet(index).max_column
        return max_column

    def get_row_data(self, index, row):
        row_list = []
        for i in self.get_sheet(index)[row]:
            row_list.append(i.value)
        return row_list


if __name__ == '__main__':
    hand_excel = HandExcel()
    hand_excel.get_sheet(0)
    print(hand_excel.get_row_data(0, 4))















