'''
第一步：加载表格文件
第二步：获取所有的sheetname
第三步：获取第一个sheet
第四步：获取某个单元格的数据
第五步：获取最大的行数
第六步：获取最大的列数
第七步：加载某一列的数据
'''
import openpyxl

class HandExcel(object):
    def load_excel(self, file_path):
        file = openpyxl.load_workbook(file_path)
        return file
    def get_sheet(self, file_path, index):
        sheet_name = self.load_excel(file_path).sheetnames
        sheet = self.load_excel(file_path)[sheet_name[index]]
        return sheet
    def get_one_value(self, file_path, index, row, col):
        value = self.get_sheet(file_path, index).cell(row, col).value
        return value
    def get_max_row(self, file_path, index):
        row = self.get_sheet(file_path, index).max_row
        return row
    def get_row_value(self, file_path, index, row):
        row_list = []
        for i in self.get_sheet(file_path, index)[row]:
            row_list.append(i.value)
        return row_list

if __name__ == '__main__':
    hand_excel = HandExcel()
    value = hand_excel.get_row_value('/Users/guojun/Desktop/驾照数据表.xlsx', 0, 3)
    print(value)